from datetime import date, datetime

from django.contrib import admin
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.http import HttpResponseRedirect, JsonResponse
from django.db.models import Q
from django.db.models import QuerySet
from django.shortcuts import render
from django.urls import path, reverse
from django.utils.html import format_html

from .email_utils import send_password_setup_email
from .forms import EmployeeProfileAdminForm
from .models import EmployeeProfile
from .services import sync_user_role
from teams.models import Team
from teams.models import TeamMember


@admin.register(EmployeeProfile)
class EmployeeProfileAdmin(admin.ModelAdmin):
    form = EmployeeProfileAdminForm
    change_form_outer_before_template = "admin/accounts/employeeprofile/before_form.html"

    list_display = (
        "employee_card",
        "department",
        "designation",
        "role_badge",
        "status_badge",
        "salary_band",
        "reporting_manager",
    )
    list_filter = (
        "department",
        "employment_type",
        "employment_status",
        "role",
        "payment_mode",
    )
    search_fields = (
        "full_name",
        "employee_code",
        "cnic",
        "official_email",
        "personal_phone",
        "user__username",
    )
    ordering = ("full_name",)
    readonly_fields = (
        "profile_photo_preview",
        "document_status",
        "compensation_summary",
        "scheduled_checkin",
        "scheduled_checkout",
        "created_by",
        "created_at",
        "updated_at",
    )
    autocomplete_fields = ("reporting_manager", "shift_master")

    fieldsets = (
        (
            "Basic Info",
            {
                "classes": ["tab"],
                "fields": (
                    ("full_name", "username"),
                    ("first_name", "last_name"),
                    ("profile_photo_preview", "profile_photo"),
                    "compensation_summary",
                    ("gender", "date_of_birth", "marital_status"),
                ),
                "description": "Core identity and login details used across the HR workspace.",
            },
        ),
        (
            "Contact",
            {
                "classes": ["tab"],
                "fields": (
                    ("personal_email", "official_email"),
                    ("personal_phone", "country"),
                    ("emergency_contact_name", "emergency_contact_number"),
                    ("city", "work_location"),
                    "current_address",
                ),
                "description": "Personal, work, and emergency contact details for quick reference.",
            },
        ),
        (
            "Employment",
            {
                "classes": ["tab"],
                "fields": (
                    ("employee_code", "department"),
                    ("designation", "team"),
                    ("employment_type", "employment_status"),
                    ("joining_date", "hire_date", "confirmation_date"),
                    ("reporting_manager", "shift_master"),
                    ("scheduled_checkin", "scheduled_checkout"),
                ),
                "description": "Department, employment status, and reporting structure.",
            },
        ),
        (
            "Identity",
            {
                "classes": ["tab"],
                "fields": (("cnic", "passport_number"),),
                "description": "Official identity details used for onboarding documentation.",
            },
        ),
        (
            "System Access",
            {
                "classes": ["tab"],
                "fields": (
                    ("role", "password"),
                    "send_setup_email",
                    ("basic_salary", "payment_mode"),
                    ("house_rent_allowance", "conveyance_allowance", "medical_allowance"),
                    "other_allowance",
                    ("bank_name", "account_title"),
                    ("account_number", "iban"),
                ),
                "description": (
                    "Managers and approvers receive staff access. "
                    "Employees stay outside the admin by default."
                ),
            },
        ),
        (
            "Documents",
            {
                "classes": ["tab"],
                "fields": (
                    "document_status",
                    ("cnic_copy", "resume"),
                    ("offer_letter", "contract"),
                ),
                "description": "Upload employee documents for record keeping.",
            },
        ),
        (
            "Notes",
            {
                "classes": ["tab"],
                "fields": (
                    "remarks",
                    ("created_by", "created_at"),
                    "updated_at",
                ),
                "description": "Internal admin notes — not visible to the employee.",
            },
        ),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "department-default-manager/",
                self.admin_site.admin_view(self.department_default_manager_view),
                name="accounts_employeeprofile_department_default_manager",
            ),
            path(
                "import-employees/",
                self.admin_site.admin_view(self.import_employees_view),
                name="accounts_employeeprofile_import_employees",
            ),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        search = request.GET.get("q", "").strip()
        status_filter = request.GET.get("status", "")
        active_tab = request.GET.get("tab", "employees")

        base_qs = self.get_queryset(request).select_related("reporting_manager")
        employees = base_qs

        if search:
            employees = employees.filter(
                Q(full_name__icontains=search)
                | Q(designation__icontains=search)
                | Q(department__icontains=search)
                | Q(employee_code__icontains=search)
                | Q(official_email__icontains=search)
            )

        if status_filter:
            employees = employees.filter(employment_status=status_filter)

        context = {
            **self.admin_site.each_context(request),
            "employees": employees,
            "total_count": base_qs.count(),
            "active_count": base_qs.filter(employment_status="Active").count(),
            "inactive_count": base_qs.filter(employment_status="Inactive").count(),
            "on_leave_count": base_qs.filter(employment_status="On Leave").count(),
            "search": search,
            "status_filter": status_filter,
            "active_tab": active_tab,
            "has_add_permission": self.has_add_permission(request),
            "opts": self.model._meta,
        }
        return render(request, "accounts/employee_directory.html", context)

    def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):
        import json
        try:
            from salary.models import TaxYear
            tax_year = TaxYear.objects.filter(is_active=True).first()
            slabs = []
            if tax_year:
                for s in tax_year.slabs.order_by("sort_order", "annual_min_income"):
                    slabs.append({
                        "name": s.name,
                        "annual_min": float(s.annual_min_income),
                        "annual_max": float(s.annual_max_income) if s.annual_max_income else None,
                        "base_tax": float(s.base_tax),
                        "rate": float(s.rate_percent),
                        "excess_over": float(s.taxable_excess_over),
                    })
            raw = json.dumps(slabs)
            context["tax_slabs_json"] = raw.replace("&", "\\u0026").replace("<", "\\u003c").replace(">", "\\u003e")
            context["tax_year_label"] = str(tax_year) if tax_year else ""
        except Exception:
            context["tax_slabs_json"] = "[]"
            context["tax_year_label"] = ""
        return super().render_change_form(request, context, add=add, change=change, form_url=form_url, obj=obj)

    def response_change(self, request, obj):
        if "_save" in request.POST:
            self.message_user(request, f'"{obj.full_name}" saved successfully.', messages.SUCCESS)
            return HttpResponseRedirect(
                reverse(
                    f"admin:{obj._meta.app_label}_{obj._meta.model_name}_change",
                    args=[obj.pk],
                )
            )
        return super().response_change(request, obj)

    def response_add(self, request, obj, post_url_continue=None):
        if "_save" in request.POST:
            self.message_user(request, f'"{obj.full_name}" created successfully.', messages.SUCCESS)
            return HttpResponseRedirect(
                reverse(
                    f"admin:{obj._meta.app_label}_{obj._meta.model_name}_change",
                    args=[obj.pk],
                )
            )
        return super().response_add(request, obj, post_url_continue)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        parent_request = request

        class RequestAwareForm(form):
            def __init__(self, *args, **inner_kwargs):
                inner_kwargs["request"] = parent_request
                super().__init__(*args, **inner_kwargs)

        return RequestAwareForm

    def import_employees_view(self, request):
        if not request.user.is_superuser:
            raise PermissionDenied

        if request.method != "POST":
            context = {**self.admin_site.each_context(request), "show_results": False}
            return render(request, "accounts/import_employees.html", context)

        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Please select an Excel file.")
            context = {**self.admin_site.each_context(request), "show_results": False}
            return render(request, "accounts/import_employees.html", context)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as exc:
            messages.error(request, f"Could not read file: {exc}")
            context = {**self.admin_site.each_context(request), "show_results": False}
            return render(request, "accounts/import_employees.html", context)

        EMPLOYMENT_TYPE_MAP = {
            "permanent (confirmed)": "Permanent",
            "permanent": "Permanent",
            "director/owner": "Permanent",
            "probation": "Probation",
            "contract": "Contract",
            "intern": "Intern",
            "resigned": "Permanent",
            "terminated": "Permanent",
            "other": "Contract",
        }
        STATUS_MAP = {"active": "Active", "inactive": "Inactive"}
        GENDER_MAP = {"m": "Male", "f": "Female", "male": "Male", "female": "Female"}
        PAYMENT_MAP = {"bank transfer": "Bank", "bank": "Bank", "cash": "Cash"}
        MARITAL_MAP = {
            "single": "Single", "married": "Married",
            "divorced": "Divorced", "widowed": "Widowed",
        }

        def _str(val):
            s = str(val).strip() if val else ""
            return "" if s in ("-", "N/A", "n/a") else s

        def _date(val):
            if not val:
                return None
            if isinstance(val, (datetime, date)):
                return val.date() if isinstance(val, datetime) else val
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(str(val).strip(), fmt).date()
                except ValueError:
                    pass
            return None

        def _decimal(val):
            try:
                return float(str(val).replace(",", "")) if val else 0
            except (ValueError, TypeError):
                return 0

        headers = [cell.value for cell in ws[1]]
        rows_data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

        created_list, updated_list, errors_list = [], [], []
        code_to_profile = {}

        for idx, d in enumerate(rows_data, start=2):
            emp_code = _str(d.get("EmployeeCode")) or _str(d.get("EmployeeID"))
            if not emp_code:
                errors_list.append({"row": idx, "code": "—", "reason": "Missing EmployeeCode"})
                continue

            first_name = _str(d.get("FirstName"))
            last_name = _str(d.get("LastName"))
            full_name = f"{first_name} {last_name}".strip() or emp_code

            official_email = _str(d.get("Email")) or f"{emp_code.lower()}@digilatics.local"
            cnic_val = _str(d.get("CNIC")) or f"IMPORT-{emp_code}"
            phone = _str(d.get("WorkPhone")) or _str(d.get("HomePhone")) or "-"

            employment_type = EMPLOYMENT_TYPE_MAP.get(_str(d.get("EmployeementType")).lower(), "Permanent")
            employment_status = STATUS_MAP.get(_str(d.get("EmployeeStatus")).lower(), "Active")
            gender = GENDER_MAP.get(_str(d.get("Gender")).lower(), "")
            marital_status = MARITAL_MAP.get(_str(d.get("MaritalStatus")).lower(), "")
            payment_mode = PAYMENT_MAP.get(_str(d.get("SalaryPaymentMethod")).lower(), "Bank")

            joining_date = _date(d.get("JoiningDate")) or date.today()
            hire_date = _date(d.get("HireDate"))
            confirmation_date = _date(d.get("ConfirmationDate"))
            dob = _date(d.get("DOB"))
            basic_salary = _decimal(d.get("BasicSalary"))

            try:
                with transaction.atomic():
                    existing_profile = EmployeeProfile.objects.filter(employee_code=emp_code).first()
                    is_update = existing_profile is not None

                    if is_update:
                        profile = existing_profile
                        user = profile.user
                        user.first_name = first_name
                        user.last_name = last_name
                        user.email = official_email
                        user.save(update_fields=["first_name", "last_name", "email"])
                    else:
                        user, _ = User.objects.get_or_create(
                            username=emp_code,
                            defaults={"email": official_email, "first_name": first_name, "last_name": last_name},
                        )
                        if not user.has_usable_password():
                            user.set_password("Change@123")
                        user.first_name = first_name
                        user.last_name = last_name
                        user.email = official_email
                        user.save()
                        # Signal auto-creates profile — get it
                        profile = EmployeeProfile.objects.get(user=user)

                    profile.full_name = full_name
                    profile.employee_code = emp_code
                    profile.gender = gender
                    profile.date_of_birth = dob
                    profile.marital_status = marital_status
                    profile.official_email = official_email
                    profile.personal_phone = phone
                    profile.emergency_contact_name = _str(d.get("EmergencyContactPerson"))
                    profile.emergency_contact_number = _str(d.get("EmergencyContactNo"))
                    profile.city = _str(d.get("City"))
                    profile.country = _str(d.get("Country"))
                    profile.current_address = _str(d.get("PermanentAddress")) or _str(d.get("TemporaryAddress"))
                    profile.cnic = cnic_val
                    profile.passport_number = _str(d.get("PassportNumber"))
                    profile.department = _str(d.get("Department"))
                    profile.designation = _str(d.get("Designation"))
                    profile.employment_type = employment_type
                    profile.employment_status = employment_status
                    profile.joining_date = joining_date
                    profile.hire_date = hire_date
                    profile.confirmation_date = confirmation_date
                    profile.basic_salary = basic_salary
                    profile.payment_mode = payment_mode
                    profile.bank_name = _str(d.get("BankName"))
                    profile.account_number = _str(d.get("AccountNumber"))
                    profile.work_location = _str(d.get("Location"))
                    profile.shift = _str(d.get("ShiftName"))
                    if not profile.created_by_id:
                        profile.created_by = request.user
                    profile.save()

                    code_to_profile[emp_code] = profile
                    entry = {"row": idx, "code": emp_code, "name": full_name}
                    (updated_list if is_update else created_list).append(entry)

            except Exception as exc:
                errors_list.append({"row": idx, "code": emp_code, "reason": str(exc)})

        # Second pass: wire up reporting managers
        for d in rows_data:
            emp_code = _str(d.get("EmployeeCode")) or _str(d.get("EmployeeID"))
            mgr_code = _str(d.get("LineManagerCode"))
            if emp_code and mgr_code and emp_code in code_to_profile:
                try:
                    mgr = EmployeeProfile.objects.get(employee_code=mgr_code)
                    p = code_to_profile[emp_code]
                    if p.reporting_manager_id != mgr.pk:
                        p.reporting_manager = mgr
                        p.save(update_fields=["reporting_manager", "updated_at"])
                except EmployeeProfile.DoesNotExist:
                    pass

        context = {
            **self.admin_site.each_context(request),
            "show_results": True,
            "created_list": created_list,
            "updated_list": updated_list,
            "errors_list": errors_list,
            "created_count": len(created_list),
            "updated_count": len(updated_list),
            "errors_count": len(errors_list),
        }
        return render(request, "accounts/import_employees.html", context)

    def department_default_manager_view(self, request):
        department = (request.GET.get("department") or "").strip()
        if not department:
            return JsonResponse({"manager_id": None, "manager_label": ""})

        team = (
            Team.objects.filter(is_active=True, department=department)
            .select_related("manager")
            .order_by("name")
            .first()
        )
        if not team or not team.manager_id:
            return JsonResponse({"manager_id": None, "manager_label": ""})

        manager_label = f"{team.manager.full_name} ({team.manager.employee_code})"
        return JsonResponse(
            {
                "manager_id": team.manager_id,
                "manager_label": manager_label,
            }
        )

    def get_queryset(self, request) -> QuerySet[EmployeeProfile]:
        queryset = (
            super()
            .get_queryset(request)
            .select_related("user", "reporting_manager", "created_by")
        )
        if request.user.is_superuser:
            return queryset

        profile = getattr(request.user, "employee_profile", None)
        if profile and profile.role == EmployeeProfile.RoleChoices.MANAGER:
            return queryset.filter(
                Q(pk=profile.pk)
                | Q(reporting_manager=profile)
                | Q(team_memberships__team__manager=profile)
            ).distinct()

        if profile and profile.role == EmployeeProfile.RoleChoices.APPROVER:
            return queryset.filter(Q(pk=profile.pk) | Q(reporting_manager=profile)).distinct()

        return queryset.none()

    def has_module_permission(self, request):
        if request.user.is_superuser:
            return True
        profile = getattr(request.user, "employee_profile", None)
        return bool(
            profile
            and profile.role in [
                EmployeeProfile.RoleChoices.MANAGER,
                EmployeeProfile.RoleChoices.APPROVER,
            ]
        )

    def has_view_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        profile = getattr(request.user, "employee_profile", None)
        if not profile:
            return False
        if profile.role in [
            EmployeeProfile.RoleChoices.MANAGER,
            EmployeeProfile.RoleChoices.APPROVER,
        ]:
            if obj is None:
                return True
            return self.get_queryset(request).filter(pk=obj.pk).exists()
        return False

    def has_add_permission(self, request):
        return request.user.is_superuser

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(
            request,
            queryset,
            search_term,
        )

        if (
            request.GET.get("app_label") == "teams"
            and request.GET.get("model_name") == "team"
            and request.GET.get("field_name") == "manager"
        ):
            queryset = queryset.filter(role=EmployeeProfile.RoleChoices.MANAGER)

        return queryset, use_distinct

    @admin.display(description="Employee")
    def employee_card(self, obj: EmployeeProfile) -> str:
        photo_html = ""
        if obj.profile_photo:
            photo_html = format_html(
                '<img src="{}" alt="{}" style="height:36px;width:36px;'
                "border-radius:12px;object-fit:cover;border:2px solid "
                'rgba(245, 132, 0, 0.2);margin-right:12px;" />',
                obj.profile_photo.url,
                obj.full_name,
            )

        return format_html(
            '<div style="display:flex;align-items:center;">{}'
            '<div><div style="font-weight:700;color:#111111;">{}</div>'
            '<div style="font-size:11px;color:#666666;">{} | {}</div></div></div>',
            photo_html,
            obj.full_name,
            obj.employee_code,
            obj.user.username,
        )

    @admin.display(description="Role")
    def role_badge(self, obj: EmployeeProfile) -> str:
        colors = {
            EmployeeProfile.RoleChoices.EMPLOYEE: ("#fff5e8", "#b45b00"),
            EmployeeProfile.RoleChoices.MANAGER: ("#111111", "#ffffff"),
            EmployeeProfile.RoleChoices.APPROVER: ("#f5f5f5", "#111111"),
            EmployeeProfile.RoleChoices.SUPER_ADMIN: ("#111111", "#ffffff"),
        }
        background, color = colors[obj.role]
        return format_html(
            '<span style="background:{};color:{};padding:6px 10px;'
            'border-radius:999px;font-size:11px;font-weight:700;">{}</span>',
            background,
            color,
            obj.role,
        )

    @admin.display(description="Status")
    def status_badge(self, obj: EmployeeProfile) -> str:
        colors = {
            EmployeeProfile.EmploymentStatusChoices.ACTIVE: ("#fff5e8", "#b45b00"),
            EmployeeProfile.EmploymentStatusChoices.ONBOARDING: ("#111111", "#ffffff"),
            EmployeeProfile.EmploymentStatusChoices.INACTIVE: ("#f5f5f5", "#525252"),
            EmployeeProfile.EmploymentStatusChoices.ON_LEAVE: ("#f0f0f0", "#111111"),
        }
        background, color = colors[obj.employment_status]
        return format_html(
            '<span style="background:{};color:{};padding:6px 10px;'
            'border-radius:999px;font-size:11px;font-weight:700;">{}</span>',
            background,
            color,
            obj.employment_status,
        )

    @admin.display(description="Monthly Cost")
    def salary_band(self, obj: EmployeeProfile) -> str:
        total = (
            obj.basic_salary
            + obj.house_rent_allowance
            + obj.conveyance_allowance
            + obj.medical_allowance
            + obj.other_allowance
        )
        return f"PKR {total:,.2f}"

    @admin.display(description="Profile Photo")
    def profile_photo_preview(self, obj: EmployeeProfile) -> str:
        if not obj or not obj.profile_photo:
            return "Upload a profile photo to preview it here."

        return format_html(
            '<img src="{}" alt="{}" style="height:88px;width:88px;'
            "border-radius:24px;object-fit:cover;border:4px solid "
            'rgba(245, 132, 0, 0.16);box-shadow:0 22px 48px rgba(17, 17, 17, 0.18);" />',
            obj.profile_photo.url,
            obj.full_name,
        )

    @admin.display(description="Documents")
    def document_status(self, obj: EmployeeProfile) -> str:
        if not obj:
            return "0 / 4 uploaded"
        return f"{obj.uploaded_documents_count} / 4 uploaded"

    @admin.display(description="Compensation Snapshot")
    def compensation_summary(self, obj: EmployeeProfile) -> str:
        if not obj:
            return "Monthly total updates after the record is saved."

        total = (
            (obj.basic_salary or 0)
            + (obj.house_rent_allowance or 0)
            + (obj.conveyance_allowance or 0)
            + (obj.medical_allowance or 0)
            + (obj.other_allowance or 0)
        )
        return format_html(
            '<div style="padding:16px 18px;border-radius:18px;background:#111111;'
            'color:#ffffff;box-shadow:0 18px 36px rgba(17,17,17,0.14);">'
            '<div style="font-size:11px;letter-spacing:.18em;text-transform:uppercase;'
            'color:rgba(255,255,255,.62);font-weight:700;">Estimated monthly cost</div>'
            '<div style="margin-top:8px;font-size:28px;font-weight:800;">PKR {:,.2f}</div>'
            '<div style="margin-top:6px;font-size:12px;color:rgba(255,255,255,.72);">'
            'Basic salary plus all configured allowances.</div></div>',
            total,
        )

    def save_model(
        self,
        request,
        obj: EmployeeProfile,
        form: EmployeeProfileAdminForm,
        change: bool,
    ) -> None:
        user = obj.user if change and obj.user_id else User()

        user.username = form.cleaned_data["username"]
        user.first_name = form.cleaned_data.get("first_name", "")
        user.last_name = form.cleaned_data.get("last_name", "")
        user.email = form.cleaned_data["official_email"]

        password = form.cleaned_data.get("password")
        send_email = form.cleaned_data.get("send_setup_email", False)

        if password:
            user.set_password(password)
        elif not change:
            # No password set — mark unusable so the setup-email link is the only way in
            user.set_unusable_password()

        user.save()

        obj.user = user

        # The post_save signal on User auto-creates a placeholder EmployeeProfile.
        # If one exists, reuse its pk so we UPDATE it instead of inserting a duplicate.
        if not change:
            try:
                existing = EmployeeProfile.objects.get(user=user)
                obj.pk = existing.pk
                obj._state.adding = False
                obj.created_at = existing.created_at
            except EmployeeProfile.DoesNotExist:
                pass

        obj.role = form.cleaned_data["role"]

        if not obj.created_by_id:
            obj.created_by = request.user

        super().save_model(request, obj, form, change)
        sync_user_role(user, obj.role, obj.employment_status)

        if not change and send_email and user.email:
            sent = send_password_setup_email(request, user)
            if sent:
                self.message_user(
                    request,
                    f"A password setup email has been sent to {user.email}.",
                    messages.SUCCESS,
                )
            else:
                self.message_user(
                    request,
                    f"Employee created, but the setup email to {user.email} could not be sent. Check your email settings.",
                    messages.WARNING,
                )

        selected_team_name = (obj.department or "").strip()

        if not selected_team_name:
            # Team cleared — remove all memberships
            TeamMember.objects.filter(employee=obj).delete()
            return

        team = (
            Team.objects.filter(is_active=True, name=selected_team_name)
            .select_related("manager")
            .first()
        )
        if not team:
            return

        # Remove memberships to any other team
        TeamMember.objects.filter(employee=obj).exclude(team=team).delete()

        if not obj.reporting_manager_id and team.manager_id:
            obj.reporting_manager = team.manager
            obj.save(update_fields=["reporting_manager", "updated_at"])

        joined_at = obj.joining_date or date.today()
        membership, _ = TeamMember.objects.get_or_create(
            team=team,
            employee=obj,
            defaults={
                "joined_at": joined_at,
                "is_primary": True,
            },
        )

        try:
            membership_changed = False
            if membership.joined_at != joined_at:
                membership.joined_at = joined_at
                membership_changed = True
            if not membership.is_primary:
                membership.is_primary = True
                membership_changed = True
            if membership_changed:
                membership.full_clean()
                membership.save()
        except ValidationError as exc:
            self.message_user(
                request,
                f"Employee saved, but team membership sync failed: {exc}",
                level=messages.WARNING,
            )
